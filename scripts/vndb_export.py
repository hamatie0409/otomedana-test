"""vndb_games.jsonl を SQLite と Excel に書き出す。

  python3 scripts/vndb_export.py
"""
import os, re, json, sqlite3, datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from common import DATA

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=11)
LINK_FONT = Font(name=FONT, size=11, color="0563C1", underline="single")

SCHEMA = """
DROP TABLE IF EXISTS games; DROP TABLE IF EXISTS characters; DROP TABLE IF EXISTS platforms;
CREATE TABLE games (
    vid TEXT PRIMARY KEY, title TEXT, title_latin TEXT, released TEXT,
    platforms TEXT, developers TEXT, publishers TEXT, devstatus TEXT,
    voiced TEXT, cv_status TEXT,
    languages TEXT, languages_official TEXT, lang_count INTEGER,
    length TEXT, length_minutes INTEGER, minage INTEGER, has_ero INTEGER,
    tags TEXT, tags_tech TEXT, staff TEXT, image_url TEXT, jawiki_url TEXT,
    rating REAL, votecount INTEGER, description TEXT, url TEXT
);
CREATE TABLE characters (
    id INTEGER PRIMARY KEY AUTOINCREMENT, vid TEXT REFERENCES games(vid),
    cid TEXT, name TEXT, name_latin TEXT, role TEXT, sex TEXT,
    cv TEXT, cv_latin TEXT, personality TEXT, appearance TEXT, char_role TEXT
);
DROP TABLE IF EXISTS traits; DROP TABLE IF EXISTS vn_tags; DROP TABLE IF EXISTS relations;
CREATE TABLE traits (cid TEXT, vid TEXT, name TEXT, category TEXT, trait TEXT);
CREATE TABLE vn_tags (vid TEXT REFERENCES games(vid), tag TEXT, category TEXT);
CREATE TABLE relations (vid TEXT REFERENCES games(vid), related_vid TEXT, type TEXT, official INTEGER);
CREATE TABLE platforms (vid TEXT REFERENCES games(vid), platform TEXT);
DROP TABLE IF EXISTS languages;
CREATE TABLE languages (vid TEXT REFERENCES games(vid), language TEXT, official INTEGER);
CREATE INDEX idx_c_vid ON characters(vid);
CREATE INDEX idx_c_cv  ON characters(cv);
CREATE INDEX idx_c_nm  ON characters(name);
CREATE INDEX idx_g_rel ON games(released);
CREATE INDEX idx_p_pl  ON platforms(platform);
CREATE INDEX idx_g_cv  ON games(cv_status);
CREATE INDEX idx_l_lang ON languages(language);
CREATE INDEX idx_tr_trait ON traits(trait);
CREATE INDEX idx_vt_tag ON vn_tags(tag);
CREATE INDEX idx_rel_vid ON relations(vid);
"""


def as_date(d):
    if d and re.fullmatch(r"\d{4}-\d{2}-\d{2}", d):
        return datetime.date(*(int(x) for x in d.split("-")))
    return d


def write_sheet(ws, headers, rows, widths, link_cols=(), date_cols=()):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill = HEAD_FONT, HEAD_FILL
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top")
    for col in date_cols:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            if isinstance(row[0].value, datetime.date):
                row[0].number_format = "yyyy/mm/dd"
    for col in link_cols:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            if row[0].value:
                row[0].hyperlink = row[0].value
                row[0].font = LINK_FONT
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


def main():
    recs = [json.loads(l) for l in open(os.path.join(DATA, "vndb_games.jsonl"), encoding="utf-8")]

    # ---------- SQLite ----------
    db = os.path.join(DATA, "vndb_otome.db")
    con = sqlite3.connect(db)
    con.executescript(SCHEMA)
    for r in recs:
        con.execute("INSERT INTO games VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (
            r["vid"], r.get("title"), r.get("title_latin"), r.get("released"),
            " / ".join(r["platforms"]), " / ".join(r["developers"]),
            " / ".join(r["publishers"]), r.get("devstatus"),
            r.get("voiced"), r.get("cv_status"),
            " / ".join(r.get("languages", [])), " / ".join(r.get("languages_official", [])),
            r.get("lang_count"),
            r.get("length"), r.get("length_minutes"), r.get("minage"),
            1 if r.get("has_ero") else 0,
            " / ".join(r.get("tags", [])), " / ".join(r.get("tags_tech", [])),
            " / ".join("%s: %s" % (k, "・".join(v)) for k, v in (r.get("staff") or {}).items()),
            r.get("image_url"), r.get("jawiki_url"),
            r.get("rating"),
            r.get("votecount"), r.get("description"), r["url"]))
        for c in r["characters"]:
            con.execute("INSERT INTO characters (vid,cid,name,name_latin,role,sex,cv,cv_latin,"
                        "personality,appearance,char_role) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        (r["vid"], c["cid"], c["character"], c["character_latin"],
                         c["role"], c["sex"], c["cv"], c["cv_latin"],
                         " / ".join(c.get("personality", [])),
                         " / ".join(c.get("appearance", [])),
                         " / ".join(c.get("char_role", []))))
            for t in c.get("traits", []):
                cat, name = t.split(":", 1)
                con.execute("INSERT INTO traits VALUES (?,?,?,?,?)",
                            (c["cid"], r["vid"], c["character"], cat, name))
        for p in r["platforms"]:
            con.execute("INSERT INTO platforms VALUES (?,?)", (r["vid"], p))
        off = set(r.get("languages_official", []))
        for l in r.get("languages", []):
            base = l.replace("(ファン翻訳)", "").replace("(機械翻訳)", "")
            con.execute("INSERT INTO languages VALUES (?,?,?)", (r["vid"], base, 1 if base in off else 0))
        for t in r.get("tags", []):
            con.execute("INSERT INTO vn_tags VALUES (?,?,?)", (r["vid"], t, "内容"))
        for t in r.get("tags_tech", []):
            con.execute("INSERT INTO vn_tags VALUES (?,?,?)", (r["vid"], t, "技術"))
        for rel in r.get("relations", []):
            con.execute("INSERT INTO relations VALUES (?,?,?,?)",
                        (r["vid"], rel["vid"], rel["type"], 1 if rel["official"] else 0))
    con.commit()
    con.close()

    # ---------- Excel ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "ゲーム一覧"
    write_sheet(ws,
        ["VNDB ID", "タイトル", "ローマ字", "発売日", "発売日(原文)", "機種",
         "開発", "発売元", "状態", "ボイス", "CV状態", "対応言語", "言語数",
         "プレイ時間", "実測(分)", "年齢制限", "18禁", "評価", "投票数",
         "タグ（内容）", "タグ（システム）", "スタッフ",
         "キャラクター（声優）", "キャラ数", "CV判明数",
         "パッケージ画像", "日本語Wikipedia", "VNDBページ"],
        [[r["vid"], r.get("title"), r.get("title_latin"),
          as_date(r.get("released")), r.get("released"),
          " / ".join(r["platforms"]), " / ".join(r["developers"]),
          " / ".join(r["publishers"]), r.get("devstatus"),
          r.get("voiced"), r.get("cv_status"),
          " / ".join(r.get("languages", [])), r.get("lang_count"),
          r.get("length"), r.get("length_minutes"),
          ("%s歳以上" % r["minage"]) if r.get("minage") else None,
          "18禁あり" if r.get("has_ero") else None,
          r.get("rating"), r.get("votecount"),
          " / ".join(r.get("tags", [])[:15]),
          " / ".join(r.get("tags_tech", [])[:12]),
          " / ".join("%s: %s" % (k, "・".join(v[:3])) for k, v in (r.get("staff") or {}).items()),
          "／".join("%s（%s）" % (c["character"] or "?", c["cv"])
                    for c in r["characters"] if c["cv"]) or None,
          len(r["characters"]),
          sum(1 for c in r["characters"] if c["cv"]),
          r.get("image_url"), r.get("jawiki_url"), r["url"]] for r in recs],
        [10, 40, 30, 12, 12, 32, 24, 24, 9, 13, 20, 40, 8,
         18, 10, 11, 9, 8, 9, 60, 50, 50, 60, 9, 10, 34, 34, 26],
        link_cols=(26, 27, 28), date_cols=(4,))

    write_sheet(wb.create_sheet("キャスト"),
        ["VNDB ID", "タイトル", "発売日", "区分", "キャラクター名", "ローマ字", "性別",
         "声優", "声優(ローマ字)", "性格", "外見", "役柄"],
        [[r["vid"], r.get("title"), r.get("released"), c["role"], c["character"],
          c["character_latin"], c["sex"], c["cv"], c["cv_latin"],
          " / ".join(c.get("personality", [])), " / ".join(c.get("appearance", [])),
          " / ".join(c.get("char_role", []))]
         for r in recs for c in r["characters"]],
        [10, 34, 12, 12, 20, 22, 6, 16, 22, 44, 44, 34])

    by_cv = defaultdict(list)
    for r in recs:
        for c in r["characters"]:
            if c["cv"]:
                by_cv[c["cv"]].append((r, c))
    write_sheet(wb.create_sheet("声優別"),
        ["声優", "ローマ字", "出演作品数", "演じたキャラ数", "代表作（新しい順）"],
        [[cv, pairs[0][1]["cv_latin"], len({p[0]["vid"] for p in pairs}), len(pairs),
          " / ".join(t for t in sorted({p[0].get("title") or "" for p in pairs}, reverse=True)[:6])]
         for cv, pairs in sorted(by_cv.items(), key=lambda kv: (-len({p[0]["vid"] for p in kv[1]}), kv[0]))],
        [20, 26, 11, 13, 70])

    lc, lo = defaultdict(int), defaultdict(int)
    for r in recs:
        off = set(r.get("languages_official", []))
        for l in r.get("languages", []):
            base = l.replace("(ファン翻訳)", "").replace("(機械翻訳)", "")
            lc[base] += 1
            if base in off:
                lo[base] += 1
    write_sheet(wb.create_sheet("言語別"), ["言語", "作品数", "公式", "ファン翻訳"],
                [[l, n, lo[l], n - lo[l]] for l, n in sorted(lc.items(), key=lambda kv: -kv[1])],
                [22, 10, 10, 12])

    tmap = {r["vid"]: r.get("title") for r in recs}
    rel_rows = []
    for r in recs:
        for rel in r.get("relations", []):
            rel_rows.append([r["vid"], r.get("title"), rel["type"],
                             tmap.get(rel["vid"], "(対象外の作品)"), rel["vid"],
                             "公式" if rel["official"] else "非公式",
                             "https://vndb.org/" + rel["vid"]])
    rel_rows.sort(key=lambda x: (x[1] or "", x[2]))
    write_sheet(wb.create_sheet("関連作品"),
                ["VNDB ID", "作品", "関係", "相手の作品", "相手ID", "公式", "相手のVNDBページ"],
                rel_rows, [10, 38, 14, 38, 10, 8, 26], link_cols=(7,))

    tagc = defaultdict(int)
    for r in recs:
        for t in r.get("tags", []):
            tagc[t] += 1
    write_sheet(wb.create_sheet("タグ別"), ["タグ（内容）", "作品数"],
                [[t, n] for t, n in sorted(tagc.items(), key=lambda kv: -kv[1])], [40, 10])

    trc = defaultdict(int)
    for r in recs:
        for c in r["characters"]:
            for t in c.get("traits", []):
                trc[t.replace(":", "：")] += 1
    write_sheet(wb.create_sheet("キャラ属性別"), ["属性", "キャラ数"],
                [[t, n] for t, n in sorted(trc.items(), key=lambda kv: -kv[1])], [40, 10])

    pc = defaultdict(int)
    for r in recs:
        for p in r["platforms"]:
            pc[p] += 1
    write_sheet(wb.create_sheet("機種別"), ["機種", "作品数"],
                [[p, n] for p, n in sorted(pc.items(), key=lambda kv: -kv[1])], [24, 10])

    out = os.path.join(DATA, "vndb_otome.xlsx")
    wb.save(out)
    print("games=%d / characters=%d / 声優=%d" % (
        len(recs), sum(len(r["characters"]) for r in recs), len(by_cv)))
    print("-> %s" % db)
    print("-> %s" % out)


if __name__ == "__main__":
    main()
