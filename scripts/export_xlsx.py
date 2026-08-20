"""games.jsonl を Excel (data/otomegame.xlsx) に書き出す。4シート構成。

  python3 scripts/export_xlsx.py
"""
import os, re, json, datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from common import DATA

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=11)
LINK_FONT = Font(name=FONT, size=11, color="0563C1", underline="single")


def as_date(iso, raw):
    """完全な日付ならdate型（Excelで並べ替え可）、そうでなければ原文のまま"""
    if iso and re.fullmatch(r"\d{4}-\d{2}-\d{2}", iso):
        return datetime.date(*(int(x) for x in iso.split("-")))
    return raw


def write_sheet(ws, headers, rows, widths, link_cols=(), date_cols=()):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill = HEAD_FONT, HEAD_FILL
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22

    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top")
    for col in date_cols:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            if isinstance(row[0].value, datetime.date):
                row[0].number_format = "yyyy/mm/dd"
    for col in link_cols:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            if row[0].value:
                row[0].hyperlink = row[0].value
                row[0].font = LINK_FONT

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


def main():
    recs = [json.loads(l) for l in open(os.path.join(DATA, "games.jsonl"), encoding="utf-8")]
    wb = Workbook()

    # --- 1. ゲーム一覧（1作品1行。キャラ・声優も同じ行に並べる） ---
    ws = wb.active
    ws.title = "ゲーム一覧"
    write_sheet(ws,
        ["ID", "タイトル", "種別", "機種", "シリーズ", "発売日", "発売日(原文)",
         "ジャンル", "サブジャンル", "価格", "メーカー", "キャラデザ", "シナリオ",
         "キャラクター（声優）", "キャスト数", "タグ",
         "公式サイト", "メーカーサイト", "画像URL", "ASIN", "最終更新", "ページURL"],
        [[r["id"], r["title"], r["game_type"], r["platform"], r["series"],
          as_date(r["release_date_iso"], r["release_date"]), r["release_date"],
          r["genre"], r["genre_sub"], r["price"], r["maker"],
          r["character_design"], r["scenario"],
          "／".join("%s（%s）" % (c["character"] or "?", c["cv"] or "?")
                    for c in r["characters"]) or None,
          len(r["characters"]),
          " / ".join(r["keywords"]) or None,
          r["official_site"], r["maker_site"], r["image_url"], r["asin"],
          r["last_updated"], r["url"]] for r in recs],
        [6, 38, 10, 18, 16, 12, 13, 16, 22, 9, 18, 13, 13, 70, 9, 22, 32, 32, 38, 12, 12, 28],
        link_cols=(17, 18, 19, 22), date_cols=(6,))

    # --- 2. キャスト（1キャラ1行） ---
    ws = wb.create_sheet("キャスト")
    write_sheet(ws,
        ["ゲームID", "タイトル", "シリーズ", "機種", "区分", "キャラクター名", "声優"],
        [[r["id"], r["title"], r["series"], r["platform"],
          c["section"], c["character"], c["cv"]]
         for r in recs for c in r["characters"]],
        [9, 38, 16, 18, 20, 22, 20])

    # --- 3. 声優別（出演本数の多い順） ---
    by_cv = defaultdict(list)
    for r in recs:
        for c in r["characters"]:
            if c["cv"]:
                by_cv[c["cv"]].append((r, c))
    rows = []
    for cv, pairs in sorted(by_cv.items(), key=lambda kv: (-len(kv[1]), kv[0])):
        rows.append([cv, len(pairs),
                     " / ".join(sorted({p[0]["series"] or "-" for p in pairs})),
                     " / ".join(sorted({p[1]["character"] or "-" for p in pairs}))])
    write_sheet(ws := wb.create_sheet("声優別"),
        ["声優", "出演作品数", "シリーズ", "演じたキャラクター"], rows,
        [20, 11, 40, 50])

    # --- 4. タグ ---
    write_sheet(wb.create_sheet("タグ"),
        ["ゲームID", "タイトル", "タグ"],
        [[r["id"], r["title"], k] for r in recs for k in r["keywords"]],
        [9, 38, 20])

    out = os.path.join(DATA, "otomegame.xlsx")
    wb.save(out)
    print("games=%d / cast=%d / 声優=%d / tags=%d -> %s" % (
        len(recs), sum(len(r["characters"]) for r in recs), len(by_cv),
        sum(len(r["keywords"]) for r in recs), out))


if __name__ == "__main__":
    main()
