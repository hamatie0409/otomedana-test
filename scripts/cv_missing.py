"""声優が判明していない作品を洗い出して data/cv_missing.xlsx に出す。

  python3 scripts/cv_missing.py

「ボイスなしの作品」は対象外（欠落ではなく、音声が存在しない）。
"""
import os, re, sqlite3, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from common import DATA

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=11)
LINK_FONT = Font(name=FONT, size=11, color="0563C1", underline="single")
HOME = ("Switch", "PS", "ニンテンドー", "Xbox", "ファミコン", "セガサターン", "ドリームキャスト")

HEADERS = ["VNDB ID", "タイトル", "ローマ字", "発売日", "機種区分", "機種",
           "開発", "発売元", "ボイス", "キャラ数", "投票数", "VNDBページ"]
WIDTHS = [10, 40, 32, 12, 11, 34, 24, 24, 13, 9, 8, 26]


def as_date(d):
    return datetime.date(*(int(x) for x in d.split("-"))) if d and re.fullmatch(r"\d{4}-\d{2}-\d{2}", d) else d


def kind(plat):
    p = plat or ""
    if any(k in p for k in HOME):
        return "家庭用機"
    if "Android" in p or "iOS" in p or "携帯" in p:
        return "スマホ"
    if "ブラウザ" in p:
        return "ブラウザ"
    if p:
        return "PC"
    return "不明"


def fetch(con, status):
    rows = con.execute("""
        SELECT g.vid, g.title, g.title_latin, g.released, g.platforms,
               g.developers, g.publishers, g.voiced, g.votecount, g.url,
               (SELECT COUNT(*) FROM characters c WHERE c.vid = g.vid)
        FROM games g WHERE g.cv_status = ?
    """, (status,)).fetchall()
    out = []
    for vid, t, lat, rel, plat, dev, pub, voi, vc, url, nch in rows:
        out.append([vid, t, lat, as_date(rel), kind(plat), plat, dev, pub,
                    voi, nch, vc or 0, url])
    # 家庭用機を先頭に、次に投票数の多い順
    order = {"家庭用機": 0, "PC": 1, "スマホ": 2, "ブラウザ": 3, "不明": 4}
    out.sort(key=lambda r: (order.get(r[4], 9), -r[10]))
    return out


def write_sheet(ws, rows):
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill = HEAD_FONT, HEAD_FILL
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22
    for r in rows:
        ws.append(r)
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top")
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        if isinstance(row[0].value, datetime.date):
            row[0].number_format = "yyyy/mm/dd"
    for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
        if row[0].value:
            row[0].hyperlink = row[0].value
            row[0].font = LINK_FONT
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


def main():
    con = sqlite3.connect(os.path.join(DATA, "vndb_otome.db"))
    a = fetch(con, "ボイスありだがCV未入力")
    b = fetch(con, "ボイス情報なし")

    wb = Workbook()
    ws = wb.active
    ws.title = "①ボイスあり・CV未入力"
    write_sheet(ws, a)
    write_sheet(wb.create_sheet("②ボイス情報なし"), b)

    out = os.path.join(DATA, "cv_missing.xlsx")
    wb.save(out)
    print("① ボイスありだがCV未入力: %d件（うち家庭用機 %d件）"
          % (len(a), sum(1 for r in a if r[4] == "家庭用機")))
    print("② ボイス情報なし        : %d件（うち家庭用機 %d件）"
          % (len(b), sum(1 for r in b if r[4] == "家庭用機")))
    print("-> %s" % out)
    return a, b


if __name__ == "__main__":
    main()
